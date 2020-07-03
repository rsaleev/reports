import asyncio
import functools
import os
import signal
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import aiosmtplib
import openpyxl
import pycron
import toml

import configuration.settings as cs
from utils.asyncsql import AsyncDBPool
from utils.asynclog import AsyncLogger

from setproctitle import setproctitle
from sdnotify import SystemdNotifier


class IncomingsNotifier:
    def __init__(self):
        self.__eventsignal = False
        self.__eventloop = None
        self.__dbconnector_ws = None
        self.__dbconnector_is = None
        self.__addresses = []

    @property
    def evntloop(self):
        return self.__eventloop

    @property
    def eventsignal(self):
        return self.__eventsignal

    @eventsignal.setter
    def eventsignal(self, v):
        self.__eventsignal = v

    async def _initialize(self):
        sd_notifier = SystemdNotifier()
        configuration = toml.load(cs.CONFIG_FILE)
        self.__logger = AsyncLogger(f"{cs.LOG_PATH}/reports.log")
        self.__dbconnector_is = AsyncDBPool(
            host=configuration["wisepark"]["rdbs"]["host"],
            port=configuration["wisepark"]["rdbs"]["port"],
            login=configuration["wisepark"]["rdbs"]["login"],
            password=configuration["wisepark"]["rdbs"]["password"],
            database=configuration["wisepark"]["rdbs"]["database"],
        )
        self.__dbconnector_ws = AsyncDBPool(
            host=configuration["integration"]["rdbs"]["host"],
            port=configuration["integration"]["rdbs"]["port"],
            login=configuration["integration"]["rdbs"]["login"],
            password=configuration["integration"]["rdbs"]["password"],
            database=configuration["integration"]["rdbs"]["database"],
        )

        connections_tasks = []
        connections_tasks.append(self.__dbconnector_is.connect())
        connections_tasks.append(self.__dbconnector_ws.connect())
        await asyncio.gather(*connections_tasks)
        reports_settings = toml.load(cs.REPORTS_FILE)
        self.__schedule = reports_settings["notify"]["incomings"]["schedule"]
        self.__addresses = reports_settings["notify"]["incomings"]["addresses"]
        self.__template = reports_settings["noitify"]["incomings"]["template"]
        return self

    async def _process(self):
        try:
            today = date.today()
            from_day = today - timedelta(days=today.weekday(), weeks=1)
            weeknum = from_day.isocalendar()[1]
            dates = [from_day + timedelta(days=d) for d in range(7)]
            tasks = []
            for d in dates:
                tasks.append(
                    self.__dbconnector_ws.callproc("rep_incomings", rows=1, values=[d])
                )
            data = await asyncio.gather(*tasks)
            wb = openpyxl.load_workbook(f"{cs.RESOURCES}/ampp/incomings_report.xlsx")
            ws = wb.active
            ws.title = cs.AMPP_PARKING_ADDRESS
            rows = [row for row in ws.iter_rows(min_row=2, max_row=8)]
            for row, val in zip(rows, data):
                ws.cell(column=1, row=row[0].row).value = val["DayWeek"]
                ws.cell(column=2, row=row[0].row).value = val["totalEntries"]
                ws.cell(column=3, row=row[0].row).value = val["totalExits"]
                ws.cell(column=4, row=row[0].row).value = val["totalPayments"]
                ws.cell(column=5, row=row[0].row).value = val["cashIncomings"]
                ws.cell(column=6, row=row[0].row).value = val["cashlessIncomings"]
                ws.cell(column=7, row=row[0].row).value = val["mobileIncomings"]
                ws.cell(column=8, row=row[0].row).value = val["totalIncomings"]
                ws.cell(column=9, row=row[0].row).value = val["lostTickets"]
                ws.cell(column=10, row=row[0].row).value = val["totalExemptions"]
            file = f"{cs.AMPP_PARKING_ID}_неделя_{weeknum}_доходность.xlsx"
            wb.save(file)
            configuration = toml.load(cs.CONFIG_FILE)
            settings = toml.load(cs.REPORTS_FILE)
            for address in settings["notify"]["incomings"]["addresses"]:
                # send data
                message = MIMEMultipart()
                message["From"] = cs.REPORTS_SMTP_ADDRESS
                message["To"] = address
                message["Subject"] = f"Доходность {configuration['ampp']['id']}"
                plain_text_message = MIMEText(
                    f"Неделя:{weeknum}\nМесяц:{from_day.month}\nГод:{from_day.year}\n",
                    "plain",
                    "utf-8",
                )
                message.attach(plain_text_message)
                fp = open(file, "rb")
                attachment = MIMEApplication(fp.read(), _subtype="xlsx")
                fp.close()
                filename = f"{Path(file).name}"
                attachment.add_header(
                    "Content-Disposition", "attachment", filename=filename
                )
                message.attach(attachment)
                plain_text_message = MIMEText(
                    "\nАвтоматизированная система генерации и отправки отчетов"
                )
                await aiosmtplib.send(
                    message,
                    hostname=configuration["reports"]["smtp"]["host"],
                    port=configuration["reports"]["smtp"]["port"],
                    username=configuration["reports"]["smtp"]["login"],
                    password=configuration["reports"]["smtp"]["password"],
                    use_tls=configuration["reports"]["smtp"]["use_tls"],
                    start_tls=configuration["reports"]["smtp"]["start_tls"],
                )
                os.remove(filename)
        except:
            pass

    async def _dispatch(self):
        while not self.eventsignal:
            settings = toml.load(cs.REPORTS_FILE)
            if (
                datetime.today().weekday() == 0
                and datetime.now().hour > 0
                and datetime.now().hour < 3
            ):
                await self._process()
                await asyncio.sleep(24 * 60 * 60)

    async def _signal_cleanup(self):
        await self.__logger.warning({"module": self.name, "msg": "Shutting down"})
        await self.__dbconnector_is.callproc(
            "cmiu_processes_upd", rows=0, values=[self.name, 0, 0, datetime.now()]
        )
        closing_tasks = []
        closing_tasks.append(self.__dbconnector_is.disconnect())
        closing_tasks.append(self.__amqpconnector.disconnect())
        closing_tasks.append(self.__logger.shutdown())
        await asyncio.gather(*closing_tasks, return_exceptions=True)

    async def _signal_handler(self, signal):
        # stop while loop coroutine
        self.eventsignal = True
        tasks = [
            task
            for task in asyncio.all_tasks(self.eventloop)
            if task is not asyncio.tasks.current_task()
        ]
        for t in tasks:
            t.cancel()
        asyncio.ensure_future(self._signal_cleanup())
        # perform eventloop shutdown
        try:
            self.eventloop.stop()
            self.eventloop.close()
        except:
            pass
        # close process
        os._exit(0)

    def run(self):
        # use own event loop
        self.eventloop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.eventloop)
        signals = (signal.SIGHUP, signal.SIGTERM, signal.SIGINT)
        # add signal handler to loop
        for s in signals:
            self.eventloop.add_signal_handler(
                s, functools.partial(asyncio.ensure_future, self._signal_handler(s))
            )
        # try-except statement
        try:
            self.eventloop.run_until_complete(self._initialize())
            self.eventloop.run_until_complete(self._dispatch())
        except asyncio.CancelledError:
            pass
