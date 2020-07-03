import asyncio
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import aiosmtplib
import configuration.settings as cs
from utils.asyncsql import AsyncDBPool
from datetime import datetime, date, timedelta
import openpyxl
from email.mime.text import MIMEText
from pathlib import Path
import signal
import os
import functools
from setproctitle import setproctitle
import toml


class ConsumablesNotifier:
    def __init__(self):
        self.__eventsignal = False
        self.__eventloop: object = None
        self.__logger: object = None
        self.__dbconnector_ws: object = None
        self.__dbconnector_is: object = None
        self.name = 'ConsumablesNotifier'
        self.type = 'notify'
        self.alias = 'consumables'
        self.__schedule = None

    @property
    def evntloop(self):
        return self.__eventloop

    @property
    def eventsignal(self):
        return self.__eventsignal

    @eventsignal.setter
    def eventsignal(self, v):
        self.__eventsignal = v

    async def _initialize(self):
        setproctitle('rep-consumables')
        configuration = toml.load(cs.CONFIG_FILE)
        self.__dbconnector_ws = AsyncDBPool(host=configuration['wisepark']['rdbs']['host'],
                                            port=configuration['wisepark']['rdbs']['port'],
                                            login=configuration['wisepark']['rdbs']['login'],
                                            password=configuration['wisepark']['rdbs']['password'],
                                            database=configuration['wisepark']['rdbs']['integration'])
        self.__dbconnector_is = AsyncDBPool(host=configuration['integration']['rdbs']['host'],
                                            port=configuration['integration']['rdbs']['port'],
                                            login=configuration['integration']['rdbs']['login'],
                                            password=configuration['integration']['rdbs']['password'],
                                            database=configuration['integration']['rdbs']['integration'])
        connections_tasks = []
        connections_tasks.append(self.__dbconnector_is.connect())
        connections_tasks.append(self.__dbconnector_ws.connect())
        await asyncio.gather(*connections_tasks)
        reports_settings = toml.load(cs.REPORTS_FILE)
        self.__schedule = reports_settings['notify']['consumables']['schedule']
        self.__addresses = reports_settings['notify']['consumables']['addresses']
        self.__template = reports_settings['notify']['consumables']['template']
        return self

    async def _process(self):
        try:
            today = date.today()
            
            from_day = today.replace(day=15, month=today.month-1)

            data = await self.__dbconnector_ws.callproc('rep_consumables', rows=1, values=[date])
            wb = openpyxl.load_workbook(f'{cs.RESOURCES}/ampp/{self.__template}')
            ws = wb.active
            rows = [row for row in ws.iter_rows(min_row=2, max_row=8)]
            for row, val in zip(rows, data):
                ws.cell(column=1, row=row[0].row).value = val['DayWeek']
                ws.cell(column=2, row=row[0].row).value = val['totalEntries']
                ws.cell(column=3, row=row[0].row).value = val['totalExits']
                ws.cell(column=4, row=row[0].row).value = val['totalPayments']
                ws.cell(column=5, row=row[0].row).value = val['cashIncomings']
                ws.cell(column=6, row=row[0].row).value = val['cashlessIncomings']
                ws.cell(column=7, row=row[0].row).value = val['mobileIncomings']
                ws.cell(column=8, row=row[0].row).value = val['totalIncomings']
                ws.cell(column=9, row=row[0].row).value = val['lostTickets']
                ws.cell(column=10, row=row[0].row).value = val['totalExemptions']
            file = f"{cs.AMPP_PARKING_ID}_неделя_{weeknum}_доходность.xlsx"
            wb.save(file)
            for address in self.__addresses:
                # send data
                if len(address) > 0:
                    message = MIMEMultipart()
                    message["From"] = cs.REPORTS_SMTP_ADDRESS
                    message["To"] = address
                    message["Subject"] = f'Доходность {cs.AMPP_PARKING_ID}'
                    plain_text_message = MIMEText(f"Неделя:{weeknum}\nМесяц:{from_day.month}\nГод:{from_day.year}\n", "plain", "utf-8")
                    message.attach(plain_text_message)
                    fp = open(file, 'rb')
                    attachment = MIMEApplication(fp.read(), _subtype="xlsx")
                    fp.close()
                    filename = f'{Path(file).name}'
                    attachment.add_header('Content-Disposition', 'attachment', filename=filename)
                    message.attach(attachment)
                    plain_text_message = MIMEText("\nАвтоматизированная система генерации отчетов")
                    await aiosmtplib.send(
                        message,
                        hostname=cs.REPORTS_SMTP_HOST,
                        port=cs.REPORTS_SMTP_PORT,
                        username=cs.REPORTS_SMTP_LOGIN,
                        password=cs.REPORTS_SMTP_PASSWORD,
                        use_tls=False,
                        start_tls=True
                    )
        except:
            raise

    async def _dispatch(self):
        while not self.eventsignal:
            if datetime.today().weekday() == 0 and datetime.now().hour > 0 and datetime.now().hour < 3:
                await self._process()
                await asyncio.sleep(24*60*60)

    async def _signal_cleanup(self):
        await self.__logger.warning({'module': self.name, 'msg': 'Shutting down'})
        await self.__dbconnector_is.callproc('cmiu_processes_upd', rows=0, values=[self.name, 0, 0, datetime.now()])
        closing_tasks = []
        closing_tasks.append(self.__dbconnector_is.disconnect())
        closing_tasks.append(self.__amqpconnector.disconnect())
        closing_tasks.append(self.__logger.shutdown())
        await asyncio.gather(*closing_tasks, return_exceptions=True)

    async def _signal_handler(self, signal):
        # stop while loop coroutine
        self.eventsignal = True
        tasks = [task for task in asyncio.all_tasks(self.eventloop) if task is not
                 asyncio.tasks.current_task()]
        for t in tasks:
            t.cancel()
        asyncio.ensure_future(self._signal_cleanup())
        # perform eventloop shutdown
        try:
            self.eventloop.stop()
            self.eventloop.close()
        except:
            pass
        # close process
        os._exit(0)

    def run(self):
        # use own event loop
        self.eventloop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.eventloop)
        signals = (signal.SIGHUP, signal.SIGTERM, signal.SIGINT)
        # add signal handler to loop
        for s in signals:
            self.eventloop.add_signal_handler(s, functools.partial(asyncio.ensure_future,
                                                                   self._signal_handler(s)))
        # try-except statement
        try:
            self.eventloop.run_until_complete(self._initialize())
            self.eventloop.run_until_complete(self._dispatch())
        except asyncio.CancelledError:
            pass


reporter = IncomingsSender()
reporter.run()
