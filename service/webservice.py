import codecs
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
import toml
import uvicorn
from fastapi import FastAPI, Form
from starlette.background import BackgroundTasks
from starlette.middleware.cors import CORSMiddleware
from starlette.requests import Request
from starlette.responses import (
    FileResponse,
    RedirectResponse,
    Response,
    StreamingResponse,
)
from starlette.staticfiles import StaticFiles
from starlette.templating import Jinja2Templates

import configuration.settings as cs
from utils.asynclog import AsyncLogger
from utils.asyncsql import AsyncDBPool

app = FastAPI()

name = "reports"


CONFIGURATION = toml.load(cs.CONFIG_FILE)
TEMPLATES = Jinja2Templates(directory=f"{Path(Path(__file__).parents[0])}/templates")
RESOURCES_DIR = f"{Path(cs.RESOURCES_DIR)}/ampp/"
DBCONNECTOR_WS = AsyncDBPool(
    hots=CONFIGURATION["wisepark"]["rdbs"]["host"],
    port=CONFIGURATION["wisepark"]["rdbs"]["port"],
    login=CONFIGURATION["wisepark"]["rdbs"]["login"],
    password=CONFIGURATION["wisepark"]["rdbs"]["password"],
    database=CONFIGURATION["wisepark"]["rdbs"]["databse"],
)


enabled = False
report = ""

TEMPLATES.env.filters["enabled"] = enabled
TEMPLATES.env.filters["option"] = report


def convert_keys(d):
    dict_converted = {}
    dict_converted["A"] = d["idx"]
    dict_converted["B"] = d["parkingId"]
    dict_converted["C"] = d["ticketNumber"]
    dict_converted["D"] = d["sessionNumber"]
    dict_converted["E"] = d["traEntryTS"]
    dict_converted["F"] = d["traPayTS"]
    dict_converted["G"] = d["traExitTS"]
    dict_converted["H"] = d["sessionDuration"]
    dict_converted["I"] = d["traPlate"]
    dict_converted["J"] = d["traPaySum"]
    dict_converted["K"] = d["traPayPaid"]
    dict_converted["L"] = d["traPayType"]
    dict_converted["M"] = d["traPayChange"]
    dict_converted["N"] = d["ticketWithoutChange"]
    dict_converted["O"] = d["payRRN"]
    dict_converted["P"] = d["sessionStatus"]
    return dict_converted


async def detailed_report_generator(report_start_date, report_stop_date):
    data = await DBCONNECTOR_WS.callproc(
        "ampp_detailedrep_get",
        rows=-1,
        values=[AMPP_ID, report_start_date, report_stop_date],
    )
    rows_data = []
    for d in data:
        d_converted = convert_keys(d)
        rows_data.append(d_converted)
    wb = openpyxl.load_workbook(f"{RESOURCES_DIR}/detailed_report.xlsx")
    ws = wb.active
    ws.unmerge_cells("A3:P3")
    num = ws["A3"].value
    ws["A3"] = num + f"{CONFIGURATION['ampp']['id']}"
    ws.merge_cells("A3:P3")
    ws.unmerge_cells("A4:P4")
    addr = ws["A4"].value
    ws["A4"] = addr + f"{cs.AMPP}"
    ws.merge_cells("A4:P4")
    ws.unmerge_cells("A5:P5")
    period = ws["A5"].value
    ws["A5"] = f"{period} {report_start_date} - {report_stop_date}"
    ws.merge_cells("A5:P5")
    ws.unmerge_cells("A7:P7")
    date = ws["A7"].value
    ws["A7"] = f"{date}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells("A7:P7")
    for rd in rows_data:
        ws.append(rd)
    filename = f"{cs.TEMPORARY_DIR}/{CONFIGURATION['ampp']['id']}_detailed.xlsx"
    wb.save(filename)
    return filename


async def consolidated_report_generator(report_start_date, report_stop_date):
    data = await DBCONNECTOR.callproc(
        "ampp_consolidatedrep_get", rows=1, values=[report_start_date, report_stop_date]
    )
    wb = openpyxl.load_workbook(f"{RESOURCES_DIR}/consolidated_report.xlsx")
    ws = wb.active
    ws.unmerge_cells("A3:E3")
    num = ws["A3"].value
    ws["A3"] = num + f"{CONFIGURATION['ampp']['id']}"
    ws.merge_cells("A3:E3")
    ws.unmerge_cells("A4:E4")
    addr = ws["A4"].value
    ws["A4"] = addr + f"{CONFIGURATION['ampp']['address']}"
    ws.merge_cells("A4:E4")
    ws.unmerge_cells("A5:E5")
    period = ws["A5"].value
    ws["A5"] = f"{period} {report_start_date} - {report_stop_date}"
    ws.merge_cells("A5:E5")
    ws.unmerge_cells("A7:E7")
    date = ws["A7"].value
    ws["A7"] = f"{date}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells("A7:E7")
    ws.unmerge_cells("D9:E9")
    ws["D9"] = data["entries"]
    ws.merge_cells("D9:E9")
    ws.unmerge_cells("D10:E10")
    ws["D10"] = data["payments"]
    ws.merge_cells("D10:E10")
    ws.unmerge_cells("D11:E11")
    ws["D11"] = data["exits"]
    ws.merge_cells("D11:E11")
    ws.unmerge_cells("D12:E12")
    ws["D12"] = data["unpaidExits"]
    ws.merge_cells("D12:E12")
    ws.unmerge_cells("D14:E14")
    ws["D14"] = data["lostTickets"]
    ws.merge_cells("D14:E14")
    ws.unmerge_cells("D15:E15")
    ws["D15"] = data["lostTicketSum"]
    ws.merge_cells("D15:E15")
    ws.unmerge_cells("D17:E17")
    ws["D17"] = data["totalPayments"]
    ws.merge_cells("D17:E17")
    ws.unmerge_cells("D18:E18")
    ws["D18"] = data["cashPayments"]
    ws.merge_cells("D18:E18")
    ws.unmerge_cells("D19:E19")
    ws["D19"] = data["cardPayments"]
    ws.merge_cells("D19:E19")
    ws.unmerge_cells("D20:E20")
    ws["D20"] = data["troikaPayments"]
    ws.merge_cells("D20:E20")
    ws.unmerge_cells("D21:E21")
    ws["D21"] = data["otherPayments"]
    ws.merge_cells("D21:E21")
    filename = f"{TEMPORARY_DIR}/{CONFIGURATION['ampp']['id']}_consolidated.xlsx"
    wb.save(filename)
    return filename


@app.on_event("startup")
async def startup():
    await DBCONNECTOR.connect()


@app.get("/")
async def homepage(request: Request):
    return TEMPLATES.TemplateResponse("index.html", {"request": request})


@app.post("/")
async def generate_report(
    request: Request,
    option: str = Form(...),
    date_from: str = Form(...),
    date_to: str = Form(...),
):
    report = option
    start_date = datetime.strptime(date_from, "%d.%m.%Y")
    end_date = datetime.strptime(date_to, "%d.%m.%Y")
    start_date_dt = start_date.strftime("%Y-%m-%d 00:00:00")
    end_date_dt = end_date.strftime("%Y-%m-%d 00:00:00")
    if option == "detailed":
        enabled = True
        filename = await detailed_report_generator(start_date_dt, end_date_dt)
    elif option == "consolidated":
        enabled = True
        filename = await consolidated_report_generator(start_date_dt, end_date_dt)
    return TEMPLATES.TemplateResponse(
        "index.html", {"request": request, "enabled": enabled}
    )


async def return_homepage(request):
    await homepage()


@app.get("/download")
async def download_report(request: Request):
    tasks = BackgroundTasks()
    for file in os.listdir(TEMPORARY_DIR):
        if file.endswith(".xlsx"):
            basename = f"{TEMPORARY_DIR}/{file}"
            tasks.add_task(os.remove, basename)
            enabled = False
            return FileResponse(basename, filename=file, background=tasks)


def run():
    uvicorn.run(
        app,
        host=cs.REPORTS_WEBSERVICE_HOST,
        port=cs.REPORTS_WEBSERVICE_PORT,
        workers=cs.REPORTS_WEBSERVICE_WORKERS,
        log_level=cs.REPORTS_WEBSERVICE_LOG_LEVEL,
    )
