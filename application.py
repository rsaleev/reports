from fastapi import FastAPI, Form
from starlette.requests import Request
from starlette.staticfiles import StaticFiles
from starlette.templating import Jinja2Templates
from starlette.responses import FileResponse, StreamingResponse, Response, RedirectResponse
import tempfile
import uvicorn
from utils.asyncsql import AsyncDBPool
import configuration as cfg
from utils.asynclog import AsyncLogger
import openpyxl
from datetime import datetime
from starlette.background import BackgroundTasks
import os
from starlette.middleware.cors import CORSMiddleware
import codecs
import re

app = FastAPI()


templates = Jinja2Templates(directory=cfg.TEMPLATES)

enabled = False
report = ''
templates.env.filters['enabled'] = enabled
templates.env.filters['option'] = report

name = "reports"

dbconnector = AsyncDBPool(cfg.wp_cnx, None)


def convert_keys(d):
    dict_converted = {}
    dict_converted['A'] = d['idx']
    dict_converted['B'] = d['parkingId']
    dict_converted['C'] = d['ticketNumber']
    dict_converted['D'] = d['sessionNumber']
    dict_converted['E'] = d['traEntryTS']
    dict_converted['F'] = d['traPayTS']
    dict_converted['G'] = d['traExitTS']
    dict_converted['H'] = d['sessionDuration']
    dict_converted['I'] = d['traPlate']
    dict_converted['J'] = d['traPaySum']
    dict_converted['K'] = d['traPayPaid']
    dict_converted['L'] = d['traPayType']
    dict_converted['M'] = d['traPayChange']
    dict_converted['N'] = d['ticketWithoutChange']
    dict_converted['O'] = d['payRRN']
    dict_converted['P'] = d['sessionStatus']
    return dict_converted


async def detailed_report_generator(report_start_date, report_stop_date):
    data = await dbconnector.callproc('ampp_detailedrep_get', rows=-1, values=[cfg.parking_id, report_start_date, report_stop_date])
    rows_data = []
    for d in data:
        d_converted = convert_keys(d)
        rows_data.append(d_converted)
    wb = openpyxl.load_workbook(f'{cfg.resources}/detailed_report.xlsx')
    ws = wb.active
    ws.unmerge_cells('A3:P3')
    num = ws['A3'].value
    ws['A3'] = num+f' {cfg.parking_id}'
    ws.merge_cells('A3:P3')
    ws.unmerge_cells('A4:P4')
    addr = ws['A4'].value
    ws['A4'] = addr+f' {cfg.parking_addr}'
    ws.merge_cells('A4:P4')
    ws.unmerge_cells('A5:P5')
    period = ws['A5'].value
    ws['A5'] = f'{period} {report_start_date} - {report_stop_date}'
    ws.merge_cells('A5:P5')
    ws.unmerge_cells('A7:P7')
    date = ws['A7'].value
    ws['A7'] = f"{date}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A7:P7')
    for rd in rows_data:
        ws.append(rd)
    filename = f'{cfg.temp}/{cfg.parking_id}_detailed.xlsx'
    wb.save(filename)
    return filename


async def consolidated_report_generator(report_start_date, report_stop_date):
    data = await dbconnector.callproc('ampp_consolidatedrep_get', rows=1, values=[report_start_date, report_stop_date])
    wb = openpyxl.load_workbook(f'{cfg.resources}/consolidated_report.xlsx')
    ws = wb.active
    ws.unmerge_cells('A3:E3')
    num = ws['A3'].value
    ws['A3'] = num+f' {cfg.parking_id}'
    ws.merge_cells('A3:E3')
    ws.unmerge_cells('A4:E4')
    addr = ws['A4'].value
    ws['A4'] = addr+f' {cfg.parking_addr}'
    ws.merge_cells('A4:E4')
    ws.unmerge_cells('A5:E5')
    period = ws['A5'].value
    ws['A5'] = f'{period} {report_start_date} - {report_stop_date}'
    ws.merge_cells('A5:E5')
    ws.unmerge_cells('A7:E7')
    date = ws['A7'].value
    ws['A7'] = f"{date}{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A7:E7')
    ws.unmerge_cells('D9:E9')
    ws['D9'] = data['entries']
    ws.merge_cells('D9:E9')
    ws.unmerge_cells('D10:E10')
    ws['D10'] = data['payments']
    ws.merge_cells('D10:E10')
    ws.unmerge_cells('D11:E11')
    ws['D11'] = data['exits']
    ws.merge_cells('D11:E11')
    ws.unmerge_cells('D12:E12')
    ws['D12'] = data['unpaidExits']
    ws.merge_cells('D12:E12')
    ws.unmerge_cells('D14:E14')
    ws['D14'] = data['lostTickets']
    ws.merge_cells('D14:E14')
    ws.unmerge_cells('D15:E15')
    ws['D15'] = data['lostTicketSum']
    ws.merge_cells('D15:E15')
    ws.unmerge_cells('D17:E17')
    ws['D17'] = data['totalPayments']
    ws.merge_cells('D17:E17')
    ws.unmerge_cells('D18:E18')
    ws['D18'] = data['cashPayments']
    ws.merge_cells('D18:E18')
    ws.unmerge_cells('D19:E19')
    ws['D19'] = data['cardPayments']
    ws.merge_cells('D19:E19')
    ws.unmerge_cells('D20:E20')
    ws['D20'] = data['troikaPayments']
    ws.merge_cells('D20:E20')
    ws.unmerge_cells('D21:E21')
    ws['D21'] = data['otherPayments']
    ws.merge_cells('D21:E21')
    filename = f'{cfg.temp}/{cfg.parking_id}_consolidated.xlsx'
    wb.save(filename)
    return filename


@app.on_event('startup')
async def startup():
    await dbconnector.connect()


@app.get("/")
async def homepage(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/")
async def generate_report(request: Request, option: str = Form(...), date_from: str = Form(...), date_to: str = Form(...)):
    report = option
    start_date = datetime.strptime(date_from, '%d.%m.%Y')
    end_date = datetime.strptime(date_to, '%d.%m.%Y')
    start_date_dt = start_date.strftime('%Y-%m-%d 00:00:00')
    end_date_dt = end_date.strftime('%Y-%m-%d 00:00:00')
    if option == 'detailed':
        enabled = True
        filename = await detailed_report_generator(start_date_dt, end_date_dt)
    elif option == 'consolidated':
        enabled = True
        filename = await consolidated_report_generator(start_date_dt, end_date_dt)
    return templates.TemplateResponse("index.html", {"request": request, "enabled": enabled})


async def return_homepage(request):
    await homepage()


@app.get("/download")
async def download_report(request: Request):
    tasks = BackgroundTasks()
    for file in os.listdir(cfg.temp):
        if file.endswith('.xlsx'):
            basename = f'{cfg.temp}/{file}'
            tasks.add_task(os.remove, basename)
            enabled = False
            return FileResponse(basename, filename=file, background=tasks)


uvicorn.run(app, host=cfg.asgi_host, port=cfg.asgi_port, workers=cfg.asgi_workers)
